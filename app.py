import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import os
from streamlit_autorefresh import st_autorefresh

def reset():
    new_workbook = load_workbook("draft_data_original.xlsx")    

    new_workbook.save("draft_data.xlsx")

    return

def update_available_players_sheet(df):

    with pd.ExcelWriter("draft_data.xlsx", mode='a', engine='openpyxl', header = False, if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name="available_players", index=False)

    return

def get_available_players():
    # Read data
    df = pd.read_excel("draft_data.xlsx", sheet_name="available_players")

    return df

def get_draft_order():
    order = pd.read_excel("draft_data.xlsx", sheet_name="draft_order")
    
    if 'Order' in order.columns:
        return order['Order'].tolist()
    else:
        return None

def get_big_board():
    order = get_draft_order()
    print(order)
    if order:
        teams = order
    else:
        teams = get_teams()

    big_board = pd.read_excel("draft_data.xlsx", sheet_name="big_board")

    for team in teams:
        p = pd.read_excel("draft_data.xlsx", sheet_name=team)
        p.rename(columns={'Player':team}, inplace=True)
        big_board = pd.merge(how = 'outer', left=big_board, right = p[['Round',team]], left_on="Round", right_on="Round")
    
    return big_board

def get_team_players(team):
    df = pd.read_excel("draft_data.xlsx", sheet_name=team)

    if not "Grade" in df.columns:
        return pd.DataFrame(columns = ["Grade", "MW", "Player"])
    else:
        return df

def get_teams():
    teams = workbook.sheetnames
    return teams[3:]

def get_row_num(sheet):
    for row_num in range(1, sheet.max_row + 2):  # +2 to check beyond max_row
        cell_value = sheet.cell(row=row_num, column=2).value # column 2 is 'B'
        if cell_value is None:
            return row_num
    
    return None

import random

def randomize():
    teams = get_teams()
    random.shuffle(teams)
    draft_order = pd.DataFrame(data = {'Order':teams})

    with pd.ExcelWriter("draft_data.xlsx", mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        draft_order.to_excel(writer, sheet_name="draft_order", index=False)

    return 

def update_team(team, Grade, MW, player):
    sheet = workbook[team]
    
    row_num = get_row_num(sheet)
                
    if row_num != None:
        for col_idx, value in enumerate([Grade, MW, player], start=2):
             sheet.cell(row=row_num, column=col_idx, value=value)

        workbook.save("draft_data.xlsx")
    else:
        st.error("Error with updating team")

    return

import time

######
# st_autorefresh(interval=3000, key="data_refresh")
st.button("Reset", on_click=reset)
st.button("Randomize Draft Order", on_click=randomize)


# Create team lists and DataFrames based on the user input
workbook = load_workbook("draft_data.xlsx")

# Display big board
st.dataframe(get_big_board(), height = 300)



# Player selection and assignment
player = st.selectbox('Select player:', get_available_players()["Player"].sort_values(ascending = True).values)
team = st.selectbox('Select team:', get_teams())

if st.button('Assign player'):
    available_players = get_available_players()
    Grade = available_players.query("Player == @player")["Grade"].values[0]
    MW = available_players.query("Player == @player")["MW"].values[0]
    player_info = pd.DataFrame(data={"Grade":Grade, "MW": MW, "Player": [player]})
    print(player_info)
    update_team(team, Grade, MW, player)

    # Remove assigned player from available players DataFrame
    available_players = available_players[available_players['Player'] != player]
    update_available_players_sheet(available_players)

    st.rerun()

# Display available players table
st.dataframe(get_available_players())

# Display assigned players tables for each team
for team in get_teams():
    with st.expander(team + ' players'):
        st.dataframe(get_team_players(team))